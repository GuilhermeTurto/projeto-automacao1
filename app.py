from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time


wb = load_workbook("DadosProcessamento.xlsx")
sheet = wb.active


with open("dados_coletados.txt", "w", encoding="utf-8") as arquivo:
    arquivo.write("=" * 60 + "\n")
    arquivo.write("    RELATÓRIO DE CONSULTAS - TJSP\n")
    arquivo.write("=" * 60 + "\n\n")


driver = webdriver.Chrome()


for row in range(1, sheet.max_row + 1):
    numero_processo = sheet.cell(row=row, column=1).value

    if not numero_processo:
        continue

    print(f"🔍 Consultando processo: {numero_processo}")

    driver.get("LINK_WEB")
    time.sleep(2)

    driver.find_element(By.XPATH, "//a[@href='https://esaj.tjsp.jus.br/esaj/portal.do?servico=190090']").click()
    time.sleep(1)
    driver.find_element(By.XPATH, "//a[@href='https://esaj.tjsp.jus.br/cpopg/open.do']").click()
    time.sleep(2)

    driver.find_element(By.XPATH, '//*[@id="radioNumeroAntigo"]').click()
    time.sleep(1)
    
    campo_pesquisa = driver.find_element(By.XPATH, '//*[@name="dadosConsulta.valorConsulta"]')
    campo_pesquisa.send_keys(numero_processo)
    time.sleep(1)
    driver.find_element(By.XPATH, '//input[@id="botaoConsultarProcessos"]').click()
    time.sleep(3)

    
    try:
        classe = driver.find_element(By.XPATH, '//span[@id="labelClasseProcesso"]').text
        dados_classe = driver.find_element(By.XPATH, '//span[@id="classeProcesso"]').text
        assunto = driver.find_element(By.XPATH, '//span[@id="labelAssuntoProcesso"]').text
        dados_assunto = driver.find_element(By.XPATH, '//span[@id="assuntoProcesso"]').text
        foro = driver.find_element(By.XPATH, '//span[@id="labelForoProcesso"]').text
        dados_foro = driver.find_element(By.XPATH, '//span[@id="foroProcesso"]').text
        vara = driver.find_element(By.XPATH, '//span[@id="labelVaraProcesso"]').text
        dados_vara = driver.find_element(By.XPATH, '//span[@id="varaProcesso"]').text
        juiz = driver.find_element(By.XPATH, '//span[@id="labelJuizProcesso"]').text
        dados_juiz = driver.find_element(By.XPATH, '//span[@id="juizProcesso"]').text
        partes = driver.find_element(By.XPATH, '//h2[@class="subtitle tituloDoBloco"]').text
        dados_partes = driver.find_element(By.XPATH, '//table[@id="tablePartesPrincipais"]').text
        movimentacoes = driver.find_element(By.XPATH, '//h2[@class="subtitle tituloDoBloco"]').text
        dados_movimentacoes = driver.find_element(By.XPATH, '//tbody[@id="tabelaUltimasMovimentacoes"]').text
    except:
        print(f"⚠ Erro ao coletar dados do processo {numero_processo}")
        continue

    
    dados_coletados = [
        ("Número do Processo", numero_processo),
        ("Classe", dados_classe),
        ("Assunto", dados_assunto),
        ("Foro", dados_foro),
        ("Vara", dados_vara),
        ("Juiz", dados_juiz),
        ("Partes", dados_partes),
        ("Movimentações", dados_movimentacoes)
    ]

    
    with open("dados_coletados.txt", "a", encoding="utf-8") as arquivo:
        arquivo.write("=" * 60 + "\n")
        arquivo.write(f"📌 Processo: {numero_processo}\n")
        arquivo.write("=" * 60 + "\n")
        for linha in dados_coletados:
            arquivo.write(f"{linha[0]:<25}: {linha[1]}\n")
        arquivo.write("\n")    

print("✅ Todos os processos foram consultados e salvos no arquivo!")


driver.quit()